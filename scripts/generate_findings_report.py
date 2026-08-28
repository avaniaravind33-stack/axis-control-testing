"""
Control Testing Findings Report Generator
Creates professional findings report with management summary
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_data():
    """Load control testing data"""
    df = pd.read_csv('data/control_testing_transactions.csv')
    df['Date'] = pd.to_datetime(df['Date'])
    return df

def generate_findings(df):
    """Generate control findings from exceptions"""
    findings = []
    finding_id = 1

    # C001: Maker/Checker Approval
    c001_exceptions = df[df['C001_Maker_Checker'] != 'Pass']
    if len(c001_exceptions) > 0:
        findings.append({
            'Finding_ID': f'F{finding_id:03d}',
            'Control': 'C001 - Maker/Checker Approval',
            'Risk': 'Unauthorized Transactions',
            'Evidence': f'{len(c001_exceptions)} transactions missing independent approval',
            'Deficiency': 'Control Failure',
            'Severity': 'High',
            'Root_Cause': 'Insufficient enforcement of approval matrix',
            'Recommendation': 'Implement automated pre-approval validation in transaction system',
            'Owner': 'Operations Manager',
            'Status': 'Open',
            'Days_Open': np.random.randint(5, 45),
            'Count': len(c001_exceptions),
            'Percentage': round(len(c001_exceptions) / len(df) * 100, 2)
        })
        finding_id += 1

    # C002: Authorization Limit Compliance
    c002_exceptions = df[df['C002_Auth_Limit'] != 'Pass']
    if len(c002_exceptions) > 0:
        findings.append({
            'Finding_ID': f'F{finding_id:03d}',
            'Control': 'C002 - Authorization Limit Compliance',
            'Risk': 'Excess Authorization / Fraud',
            'Evidence': f'{len(c002_exceptions)} transactions exceeded officer authorization limits',
            'Deficiency': 'Control Weakness',
            'Severity': 'High',
            'Root_Cause': 'Authorization limits not enforced at transaction posting',
            'Recommendation': 'Strengthen maker/checker enforcement with real-time limit validation',
            'Owner': 'Compliance Manager',
            'Status': 'Open',
            'Days_Open': np.random.randint(10, 50),
            'Count': len(c002_exceptions),
            'Percentage': round(len(c002_exceptions) / len(df) * 100, 2)
        })
        finding_id += 1

    # C003: Segregation of Duties
    c003_exceptions = df[df['C003_SOD'] != 'Pass']
    if len(c003_exceptions) > 0:
        findings.append({
            'Finding_ID': f'F{finding_id:03d}',
            'Control': 'C003 - Segregation of Duties',
            'Risk': 'Override of Controls / Fraud',
            'Evidence': f'{len(c003_exceptions)} instances of SOD violation detected',
            'Deficiency': 'Design Deficiency',
            'Severity': 'Critical',
            'Root_Cause': 'User roles not properly segregated in system',
            'Recommendation': 'Implement role-based access controls (RBAC) with IT Security',
            'Owner': 'IT Security Manager',
            'Status': 'Open',
            'Days_Open': np.random.randint(15, 60),
            'Count': len(c003_exceptions),
            'Percentage': round(len(c003_exceptions) / len(df) * 100, 2)
        })
        finding_id += 1

    # C004: 4-Eyes Principle
    c004_exceptions = df[df['C004_4Eyes'] != 'Pass']
    if len(c004_exceptions) > 0:
        findings.append({
            'Finding_ID': f'F{finding_id:03d}',
            'Control': 'C004 - 4-Eyes Principle',
            'Risk': 'Unauthorized High-Value Transactions',
            'Evidence': f'{len(c004_exceptions)} high-value transactions without dual approval',
            'Deficiency': 'Control Failure',
            'Severity': 'High',
            'Root_Cause': 'Insufficient monitoring of dual-approval requirement',
            'Recommendation': 'Enforce mandatory second approval for transactions > ₹10L with system rules',
            'Owner': 'Operations Manager',
            'Status': 'Open',
            'Days_Open': np.random.randint(3, 30),
            'Count': len(c004_exceptions),
            'Percentage': round(len(c004_exceptions) / len(df) * 100, 2)
        })
        finding_id += 1

    # C005: Duplicate Detection
    c005_exceptions = df[df['C005_Duplicate'] != 'Pass']
    if len(c005_exceptions) > 0:
        findings.append({
            'Finding_ID': f'F{finding_id:03d}',
            'Control': 'C005 - Duplicate Detection',
            'Risk': 'Duplicate Processing / Financial Loss',
            'Evidence': f'{len(c005_exceptions)} potential duplicate transactions detected',
            'Deficiency': 'Control Weakness',
            'Severity': 'Medium',
            'Root_Cause': 'Duplicate detection rules not comprehensive enough',
            'Recommendation': 'Enhance duplicate detection rules to cover all transaction corridors and amounts',
            'Owner': 'Operations Manager',
            'Status': 'Open',
            'Days_Open': np.random.randint(8, 35),
            'Count': len(c005_exceptions),
            'Percentage': round(len(c005_exceptions) / len(df) * 100, 2)
        })
        finding_id += 1

    return pd.DataFrame(findings)

def create_findings_workbook(df, findings_df):
    """Create professional findings report workbook"""

    # Calculate management summary metrics
    total_tested = len(df)
    total_exceptions = len(df[df['Deficiency Class'] != 'None'])
    critical = len(df[df['Deficiency Class'] == 'Critical'])
    major = len(df[df['Deficiency Class'] == 'Major'])
    minor = len(df[df['Deficiency Class'] == 'Minor'])

    # Risk classification
    high_risk = critical + (major // 2)  # Some majors are high risk
    medium_risk = (major - major // 2) + minor
    low_risk = 0

    open_findings = len(findings_df)
    overdue = len(findings_df[findings_df['Days_Open'] > 30])

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ==================== MANAGEMENT SUMMARY ====================
    ws_summary = wb.create_sheet('Management Summary', 0)

    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    metric_font = Font(name='Calibri', size=12, bold=True)
    metric_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    value_font = Font(name='Calibri', size=14, bold=True)
    critical_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    critical_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    high_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    high_font = Font(name='Calibri', size=14, bold=True, color='000000')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Title
    ws_summary.merge_cells('A1:D1')
    cell = ws_summary['A1']
    cell.value = 'AXIS BANK - CONTROL TESTING FINDINGS REPORT'
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align
    ws_summary.row_dimensions[1].height = 30

    # Subtitle
    ws_summary.merge_cells('A2:D2')
    cell = ws_summary['A2']
    cell.value = f'Management Summary | Report Date: {datetime.now().strftime("%B %d, %Y")}'
    cell.font = Font(name='Calibri', size=10, italic=True)
    cell.alignment = center_align

    ws_summary.merge_cells('A3:D3')
    cell = ws_summary['A3']
    cell.value = f'Testing Period: January 2024 - December 2024'
    cell.font = Font(name='Calibri', size=10, italic=True)
    cell.alignment = center_align

    # Key metrics
    row = 5
    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Controls Tested'
    cell.font = metric_font
    cell.fill = metric_fill
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = 5
    cell.font = value_font
    cell.alignment = center_align

    row += 1
    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Transactions Tested'
    cell.font = metric_font
    cell.fill = metric_fill
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = total_tested
    cell.font = value_font
    cell.alignment = center_align

    row += 1
    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Exceptions Identified'
    cell.font = metric_font
    cell.fill = metric_fill
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = total_exceptions
    cell.font = value_font
    cell.alignment = center_align

    row += 1
    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Compliance Rate'
    cell.font = metric_font
    cell.fill = metric_fill
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = round((total_tested - total_exceptions) / total_tested * 100, 2)
    cell.font = value_font
    cell.alignment = center_align
    cell.number_format = '0.00"%"'

    row += 2

    # Risk summary
    ws_summary.merge_cells(f'A{row}:D{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'RISK ASSESSMENT SUMMARY'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1

    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Critical Risk (High Severity)'
    cell.font = metric_font
    cell.fill = critical_fill
    cell.font = critical_font
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = critical
    cell.font = critical_font
    cell.fill = critical_fill
    cell.alignment = center_align
    row += 1

    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'High Risk (Major Deficiencies)'
    cell.font = metric_font
    cell.fill = high_fill
    cell.font = high_font
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = major
    cell.font = high_font
    cell.fill = high_fill
    cell.alignment = center_align
    row += 1

    ws_summary.merge_cells(f'A{row}:B{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'Medium Risk (Minor Deficiencies)'
    cell.font = metric_font
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    cell.alignment = center_align

    cell = ws_summary[f'C{row}']
    cell.value = minor
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = center_align
    row += 2

    # Top findings
    ws_summary.merge_cells(f'A{row}:D{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'TOP 3 CONTROL WEAKNESSES'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1

    for idx, (_, finding) in enumerate(findings_df.nlargest(3, 'Count').iterrows(), 1):
        ws_summary.merge_cells(f'A{row}:D{row}')
        cell = ws_summary[f'A{row}']
        cell.value = f'{idx}. {finding["Control"]}: {finding["Count"]} exceptions ({finding["Percentage"]}%)'
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_summary.row_dimensions[row].height = 20
        row += 1

    row += 1

    # Recommendations
    ws_summary.merge_cells(f'A{row}:D{row}')
    cell = ws_summary[f'A{row}']
    cell.value = 'RECOMMENDED REMEDIATION ACTIONS'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1

    recommendations = [
        '1. Implement automated pre-approval validation in transaction processing system',
        '2. Strengthen maker/checker enforcement with real-time authorization limit validation',
        '3. Enhance duplicate detection rules across all corridors and transaction amounts',
        '4. Implement role-based access controls (RBAC) to ensure segregation of duties',
        '5. Establish automated monitoring dashboard for ongoing control compliance'
    ]

    for rec in recommendations:
        ws_summary.merge_cells(f'A{row}:D{row}')
        cell = ws_summary[f'A{row}']
        cell.value = rec
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_summary.row_dimensions[row].height = 20
        row += 1

    # Set column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 25

    # ==================== FINDINGS DETAIL ====================
    ws_findings = wb.create_sheet('Findings Detail', 1)

    # Header
    headers = ['Finding ID', 'Control', 'Risk', 'Evidence', 'Deficiency', 'Severity',
               'Root Cause', 'Recommendation', 'Owner', 'Status', 'Days Open']

    for col, header in enumerate(headers, 1):
        cell = ws_findings.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Data rows
    for idx, (_, finding) in enumerate(findings_df.iterrows(), 2):
        ws_findings.cell(row=idx, column=1).value = finding['Finding_ID']
        ws_findings.cell(row=idx, column=2).value = finding['Control']
        ws_findings.cell(row=idx, column=3).value = finding['Risk']
        ws_findings.cell(row=idx, column=4).value = finding['Evidence']
        ws_findings.cell(row=idx, column=5).value = finding['Deficiency']
        ws_findings.cell(row=idx, column=6).value = finding['Severity']
        ws_findings.cell(row=idx, column=7).value = finding['Root_Cause']
        ws_findings.cell(row=idx, column=8).value = finding['Recommendation']
        ws_findings.cell(row=idx, column=9).value = finding['Owner']
        ws_findings.cell(row=idx, column=10).value = finding['Status']
        ws_findings.cell(row=idx, column=11).value = finding['Days_Open']

        # Color severity
        severity_cell = ws_findings.cell(row=idx, column=6)
        if finding['Severity'] == 'Critical':
            severity_cell.fill = critical_fill
            severity_cell.font = critical_font
        elif finding['Severity'] == 'High':
            severity_cell.fill = high_fill
            severity_cell.font = high_font
        elif finding['Severity'] == 'Medium':
            severity_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            severity_cell.font = Font(name='Calibri', size=10, bold=True)

        # Format all cells
        for col in range(1, len(headers) + 1):
            cell = ws_findings.cell(row=idx, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws_findings.row_dimensions[idx].height = 25

    # Set column widths
    ws_findings.column_dimensions['A'].width = 12
    ws_findings.column_dimensions['B'].width = 20
    ws_findings.column_dimensions['C'].width = 20
    ws_findings.column_dimensions['D'].width = 30
    ws_findings.column_dimensions['E'].width = 18
    ws_findings.column_dimensions['F'].width = 12
    ws_findings.column_dimensions['G'].width = 25
    ws_findings.column_dimensions['H'].width = 30
    ws_findings.column_dimensions['I'].width = 18
    ws_findings.column_dimensions['J'].width = 12
    ws_findings.column_dimensions['K'].width = 12

    # ==================== EXCEPTION SUMMARY ====================
    ws_exceptions = wb.create_sheet('Exception Summary', 2)

    # Headers
    headers = ['Control', 'Pass', 'Fail', 'Exception Rate', 'Top Reason', 'Remediation Status']
    for col, header in enumerate(headers, 1):
        cell = ws_exceptions.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Control data
    controls_data = [
        ('C001 - Maker/Checker', len(df[df['C001_Maker_Checker'] == 'Pass']), len(df[df['C001_Maker_Checker'] != 'Pass']), 'Missing approval', 'In Progress'),
        ('C002 - Auth Limits', len(df[df['C002_Auth_Limit'] == 'Pass']), len(df[df['C002_Auth_Limit'] != 'Pass']), 'Exceeded limits', 'In Progress'),
        ('C003 - SOD', len(df[df['C003_SOD'] == 'Pass']), len(df[df['C003_SOD'] != 'Pass']), 'User segregation', 'Not Started'),
        ('C004 - 4-Eyes', len(df[df['C004_4Eyes'] == 'Pass']), len(df[df['C004_4Eyes'] != 'Pass']), 'Missing dual approval', 'In Progress'),
        ('C005 - Duplicates', len(df[df['C005_Duplicate'] == 'Pass']), len(df[df['C005_Duplicate'] != 'Pass']), 'Duplicate detection', 'In Progress'),
    ]

    for idx, (control, passed, failed, reason, status) in enumerate(controls_data, 2):
        ws_exceptions.cell(row=idx, column=1).value = control
        ws_exceptions.cell(row=idx, column=2).value = passed
        ws_exceptions.cell(row=idx, column=3).value = failed

        rate_cell = ws_exceptions.cell(row=idx, column=4)
        rate_cell.value = round(failed / len(df) * 100, 2)
        rate_cell.number_format = '0.00"%"'

        ws_exceptions.cell(row=idx, column=5).value = reason
        ws_exceptions.cell(row=idx, column=6).value = status

        # Format
        for col in range(1, len(headers) + 1):
            cell = ws_exceptions.cell(row=idx, column=col)
            cell.border = border
            cell.alignment = center_align

    ws_exceptions.column_dimensions['A'].width = 20
    ws_exceptions.column_dimensions['B'].width = 12
    ws_exceptions.column_dimensions['C'].width = 12
    ws_exceptions.column_dimensions['D'].width = 15
    ws_exceptions.column_dimensions['E'].width = 25
    ws_exceptions.column_dimensions['F'].width = 18

    return wb

def main():
    print("Loading data...")
    df = load_data()

    print("Generating findings...")
    findings_df = generate_findings(df)

    print("Creating workbook...")
    wb = create_findings_workbook(df, findings_df)

    output_path = 'excel/Control_Testing_Findings_Report.xlsx'
    wb.save(output_path)

    print(f"✓ Findings report saved to {output_path}")
    print(f"✓ {len(findings_df)} findings identified")

if __name__ == '__main__':
    main()
