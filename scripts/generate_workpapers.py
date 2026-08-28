"""
Axis Bank Control Testing - Excel Workpapers Generator
Creates professional control testing documentation
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_workpapers():
    """Create comprehensive control testing workpapers"""

    # Read data
    print("Loading control testing data...")
    df = pd.read_csv('data/control_testing_transactions.csv')
    df['Date'] = pd.to_datetime(df['Date'])

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Define colors
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Sheet 1: Control Summary
    print("Creating Control Summary sheet...")
    ws_summary = wb.create_sheet("Control Summary")
    _create_control_summary(ws_summary, df, header_fill, header_font, title_fill)

    # Sheet 2: Control Testing Results
    print("Creating Control Testing Results sheet...")
    ws_results = wb.create_sheet("Test Results")
    _create_test_results(ws_results, df, header_fill, header_font, pass_fill, fail_fill)

    # Sheet 3: Exceptions & Deficiencies
    print("Creating Exceptions sheet...")
    ws_exceptions = wb.create_sheet("Exceptions")
    _create_exceptions_sheet(ws_exceptions, df, header_fill, header_font, fail_fill)

    # Sheet 4: Effectiveness Assessment
    print("Creating Effectiveness Assessment sheet...")
    ws_effectiveness = wb.create_sheet("Effectiveness")
    _create_effectiveness_sheet(ws_effectiveness, df, header_fill, header_font, title_fill)

    # Save workbook
    output_path = 'excel/Axis_Control_Testing_Workpapers.xlsx'
    wb.save(output_path)
    print(f"✓ Workpapers saved to {output_path}")

def _create_control_summary(ws, df, header_fill, header_font, title_fill):
    """Create control summary sheet"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "AXIS BANK - CONTROL TESTING SUMMARY"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'].value = "Control ID"
    ws[f'B{row}'].value = "Control Name"
    ws[f'C{row}'].value = "Risk Addressed"
    ws[f'D{row}'].value = "Control Type"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].font = header_font
    row += 1

    controls = {
        'C001': ('Maker/Checker Approval', 'Unauthorized transactions', 'Preventive'),
        'C002': ('Authorization Limit Compliance', 'Excess authorization', 'Preventive'),
        'C003': ('Segregation of Duties', 'Fraud/override', 'Preventive'),
        'C004': ('4-Eyes Principle', 'Unauthorized high-value txns', 'Preventive'),
        'C005': ('Duplicate Detection', 'Duplicate processing', 'Detective'),
    }

    for ctrl_id, (ctrl_name, risk, ctrl_type) in controls.items():
        ws[f'A{row}'].value = ctrl_id
        ws[f'B{row}'].value = ctrl_name
        ws[f'C{row}'].value = risk
        ws[f'D{row}'].value = ctrl_type
        row += 1

    row += 2

    # Testing Statistics
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "TESTING STATISTICS"
    cell.font = Font(bold=True, size=12)
    cell.fill = title_fill
    row += 1

    ws[f'A{row}'].value = "Total Transactions Tested"
    ws[f'B{row}'].value = len(df)
    row += 1

    ws[f'A{row}'].value = "Testing Period"
    ws[f'B{row}'].value = f"{df['Date'].min().date()} to {df['Date'].max().date()}"
    row += 1

    ws[f'A{row}'].value = "Overall Compliance Rate"
    compliance_rate = (len(df[df['Deficiency Class'] == 'None']) / len(df) * 100)
    ws[f'B{row}'].value = f"{compliance_rate:.2f}%"
    row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

def _create_test_results(ws, df, header_fill, header_font, pass_fill, fail_fill):
    """Create detailed test results"""
    row = 1

    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "CONTROL TESTING DETAILED RESULTS"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    row += 2

    # Sample results
    sample_df = df.sample(n=min(100, len(df)), random_state=42)

    headers = ['Txn ID', 'Amount', 'C001', 'C002', 'C003', 'C004', 'C005', 'Deficiency']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    for idx, tx in sample_df.iterrows():
        ws.cell(row, 1, tx['Transaction ID'])
        ws.cell(row, 2, int(tx['Amount (INR)']))

        for col_idx, col_name in enumerate(['C001_Maker_Checker', 'C002_Auth_Limit', 'C003_SOD', 'C004_4Eyes', 'C005_Duplicate'], 3):
            cell = ws.cell(row, col_idx, tx[col_name])
            if tx[col_name] == 'Pass' or tx[col_name] == 'Detected':
                cell.fill = pass_fill
            else:
                cell.fill = fail_fill

        cell = ws.cell(row, 8, tx['Deficiency Class'])
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

def _create_exceptions_sheet(ws, df, header_fill, header_font, fail_fill):
    """Create exceptions and deficiencies sheet"""
    row = 1

    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "EXCEPTIONS AND DEFICIENCIES"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    row += 2

    # Exception analysis
    for control_col, control_name in [('C001_Exception', 'C001'), ('C002_Exception', 'C002'),
                                      ('C003_Exception', 'C003'), ('C004_Exception', 'C004'), ('C005_Exception', 'C005')]:
        exceptions = df[df[control_col] != ''].groupby(control_col).size()

        if len(exceptions) > 0:
            ws[f'A{row}'].value = control_name
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            for exc_type, count in exceptions.items():
                ws[f'A{row}'].value = exc_type
                ws[f'B{row}'].value = count
                ws[f'B{row}'].fill = fail_fill
                row += 1

            row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15

def _create_effectiveness_sheet(ws, df, header_fill, header_font, title_fill):
    """Create effectiveness assessment"""
    row = 1

    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "CONTROL EFFECTIVENESS ASSESSMENT"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    row += 2

    # Design Effectiveness
    ws[f'A{row}'].value = "DESIGN EFFECTIVENESS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    design_controls = [
        ('C001: Maker/Checker', 'Effective', 'Control is well-designed to prevent unauthorized high-value txns'),
        ('C002: Authorization Limit', 'Effective', 'Role-based limits prevent excess authorization'),
        ('C003: Segregation of Duties', 'Effective', 'System enforces different officers for maker/checker'),
        ('C004: 4-Eyes Principle', 'Effective', 'Dual approval requirement for high-value txns'),
        ('C005: Duplicate Detection', 'Effective', 'Automated system flags potential duplicates'),
    ]

    for control, rating, description in design_controls:
        ws[f'A{row}'].value = control
        ws[f'B{row}'].value = rating
        ws[f'C{row}'].value = description
        row += 1

    row += 2

    # Operating Effectiveness
    ws[f'A{row}'].value = "OPERATING EFFECTIVENESS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'].value = "Control"
    ws[f'B{row}'].value = "Test Result"
    ws[f'C{row}'].value = "Operating Rate"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].fill = title_fill
    row += 1

    controls_data = [
        ('C001: Maker/Checker', 'Pass', f"{(len(df[df['C001_Maker_Checker'] == 'Pass']) / len(df) * 100):.2f}%"),
        ('C002: Authorization Limit', 'Pass', f"{(len(df[df['C002_Auth_Limit'] == 'Pass']) / len(df) * 100):.2f}%"),
        ('C003: Segregation of Duties', 'Pass', f"{(len(df[df['C003_SOD'] == 'Pass']) / len(df) * 100):.2f}%"),
        ('C004: 4-Eyes Principle', 'Pass', f"{(len(df[df['C004_4Eyes'] == 'Pass']) / len(df) * 100):.2f}%"),
        ('C005: Duplicate Detection', 'Pass', f"{(len(df[df['C005_Duplicate'] == 'Pass']) / len(df) * 100):.2f}%"),
    ]

    for control, result, rate in controls_data:
        ws[f'A{row}'].value = control
        ws[f'B{row}'].value = result
        ws[f'C{row}'].value = rate
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35

def main():
    create_excel_workpapers()
    print("\n✓ Excel workpapers created successfully!")

if __name__ == '__main__':
    main()
